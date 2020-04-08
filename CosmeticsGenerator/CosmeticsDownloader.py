
import openpyxl
import requests
from pathlib import Path

apocalyptechSheet = 'https://docs.google.com/spreadsheets/d/1v-F_3C2ceaFKJae1b6wmbelw_jLjmPPriBLzGTZMqRc/export?format=xlsx&id=1v-F_3C2ceaFKJae1b6wmbelw_jLjmPPriBLzGTZMqRc'
emotes, decos, heads, skins, echos = [],[],[],[],[]

def requestGoogleSheetAsXLSX(url):
    print("Requesting google sheet: {}".format(url))
    r = requests.get(url, allow_redirects=True)
    open('sheetData.xlsx', 'wb').write(r.content)
    print("Google sheet written...")

requestGoogleSheetAsXLSX(apocalyptechSheet)

xlsxFile = Path('sheetData.xlsx')
workbook = openpyxl.load_workbook(xlsxFile)
sheetNames = workbook.sheetnames[1:]
print("Sheet Names: {}".format(sheetNames))

for sheetName in sheetNames:
    worksheet = workbook[sheetName]
    for row in range(2, worksheet.max_row + 1): # Iterate through each row of the worksheet
        cellName = "C{}".format(row)
        invBalance = worksheet[cellName].value

        if(invBalance == None): 
            break
        
        assetPath = invBalance.replace("InvBal_","")

        if ("01_Wave" in assetPath or "02_Cheer" in assetPath or "03_Point" in assetPath or "04_Laugh" in assetPath) or ("default" in assetPath.lower()): 
            continue

        if "Heads" in sheetName:
            heads += [assetPath]
        elif "Skins" in sheetName and "Weapon" not in sheetName:
            skins += [assetPath]
        elif "Emotes" in sheetName:
            emotes += [assetPath]
        elif "Themes" in sheetName:
            echos += [assetPath]
        elif "Decorations" in sheetName:
            decos += [assetPath]

    print("Done reading {}".format(sheetName))

pathFormat = "public static readonly List<string> {0}AssetPaths = new List<string>()"

assetSets = []
nameToList = {"emotes":emotes, "deco":decos, "head":heads,"skin":skins,"echo":echos}

for name in nameToList:
    list = nameToList[name]
    assetSet = pathFormat.format(name) + '{\n'
    for asset in list:
        assetSet += '"{0}",\n'.format(asset)
    assetSet = assetSet[:-2] + "\n};"
    assetSets += [assetSet]

with open('output.txt', 'w') as outFile:
    completeSet = ""

    for assetSet in assetSets:
        completeSet += (assetSet + '\n')

    outFile.write(completeSet)
